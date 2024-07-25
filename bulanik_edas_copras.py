def ucgensel_toplam(a, b):
	return (a[0] + b[0], a[1] + b[1], a[2] + b[2])

def ucgensel_cik(a, b):
	return (a[0] - b[2], a[1] - b[1], a[2] - b[0])

def ucgensel_bol(a, b):
	return (a[0] / b[2], a[1] / b[1], a[2] / b[0])

def ucgensel_carp(a, b):
	return (a[0] * b[0], a[1] * b[1], a[2] * b[2])

def ucgensel_toplam_sabit(a, b):
	return (a[0] + b, a[1] + b, a[2] + b)

def ucgensel_cik_sabit(a, b):
	return (a[0] - b, a[1] - b, a[2] - b)

def ucgensel_cik_sabit_rev(a, b):
	return (b - a[0], b - a[1],b - a[2])

def ucgensel_bol_sabit(a, b):
	if b >= 0:
		return (a[0] / b, a[1] / b, a[2] / b)
	else:
		return (a[2] / b, a[1] / b, a[0] / b)

def ucgensel_carp_sabit(a, b):
	if b >= 0:
		return (a[0] * b, a[1] * b, a[2] * b)
	else:
		return (a[2] * b, a[1] * b, a[0] * b)


def kav(a):
	return (a[0] + a[1] + a[2]) / 3.0

def psi(a):
	if kav(a) > 0:
		return a
	else:
		return (0, 0, 0)

def max2(arr):
	last_maks = 0
	maks = 0
	for i in range(len(arr)):
		last_maks = arr[i]
		if last_maks > maks:
			maks = last_maks

	return maks

def max(arr):
	last_maks = 0
	maks = 0
	for i in range(len(arr)):
		last_maks = kav(arr[i])
		if last_maks > maks:
			maks = last_maks

	return maks

def copras_i(arr, ks, cs, kss, copras):
	for i in range(cs):
		for j in range(ks):
			last_min = 0
			mini = 100
			for k in range(kss):
				last_min = arr[k][i][j][0]
				if last_min < mini:
					mini = last_min

			my_list = list(copras[i][j])
			my_list[0] = mini
			copras[i][j] = tuple(my_list)


def copras_j(arr, ks, cs, kss, copras):
	for i in range(cs):
		for j in range(ks):
			summ = 0
			for k in range(kss):
				summ += arr[k][i][j][1]

			my_list = list(copras[i][j])
			my_list[1] = summ / kss
			copras[i][j] = tuple(my_list)

def copras_k(arr, ks, cs, kss, copras):
	for i in range(cs):
		for j in range(ks):
			last_max = 0
			maks = 0
			for k in range(kss):
				last_maks = arr[k][i][j][2]
				if last_maks > maks:
					maks = last_maks

			my_list = list(copras[i][j])
			my_list[2] = maks
			copras[i][j] = tuple(my_list)

def tup_to_str(a):
	return "({:.2f},{:.2f},{:.2f})".format(a[0], a[1], a[2])

def tup_to_str_norm(a):
	return "({},{},{})".format(a[0], a[1], a[2])

def str_to_tup(a):
	b = a[1:len(a)-1]
	c = b.split(",")
	return (c[0], c[1], c[2])

def alternatif_karsilik_ucgensel_tup(a):
	if a == 1:
		return (0, 0 , 1)
	elif a == 2:
		return (0, 1, 3)
	elif a == 3:
		return (1, 3, 5)
	elif a == 4:
		return (3, 5, 7)
	elif a == 5:
		return (5, 7, 9)
	elif a == 6:
		return (7, 9, 10)
	elif a == 7:
		return (9, 10, 10)


def alternatif_karsilik_ucgensel(a):
	if a == 1:
		return tup_to_str_norm((0, 0 , 1))
	elif a == 2:
		return tup_to_str_norm((0, 1, 3))
	elif a == 3:
		return tup_to_str_norm((1, 3, 5))
	elif a == 4:
		return tup_to_str_norm((3, 5, 7))
	elif a == 5:
		return tup_to_str_norm((5, 7, 9))
	elif a == 6:
		return tup_to_str_norm((7, 9, 10))
	elif a == 7:
		return tup_to_str_norm((9, 10, 10))

def alternatif_karsilik_sozel(a):
	if a == 1:
		return "ÇD"
	elif a == 2:
		return "D"
	elif a == 3:
		return "OD"
	elif a == 4:
		return "O"
	elif a == 5:
		return "OY"
	elif a == 6:
		return "Y"
	elif a == 7:
		return "ÇY"


cozum_sayisi = 0
kriter_sayisi = 0
katilimci_sayisi = 0

import numpy as np
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

cozum_sayisi = 8
kriter_sayisi = 22
katilimci_sayisi = 11

all_tups = []
all_tups_copras = []

for i in range(katilimci_sayisi):
	lines = []
	while True:
	    line = input("Katılımcı {} :".format(i+1))
	    if line:
	        lines.append(line.replace('\t', ','))
	    else:
	        break

	text = '\n'.join(lines)
	test1 = text.split("\n")
	b = []
	for l in range(len(test1)):
		lst_int = [int(o) for o in test1[l].split(',')]
		b.append(lst_int)

	#print(b)

	arr = np.array(b)

	arr2 = np.transpose(arr).tolist()

	arr3 = np.transpose(arr).tolist()

	arr4 = np.transpose(arr).tolist()

	arr5 = np.transpose(arr).tolist()

	arr6 = arr.tolist()

	for y in range(len(arr3)):
		for x in range(len(arr3[y])):
			arr3[y][x] = alternatif_karsilik_sozel(arr3[y][x])
			arr4[y][x] = alternatif_karsilik_ucgensel(arr4[y][x])
			arr5[y][x] = alternatif_karsilik_ucgensel_tup(arr5[y][x])

	for y in range(len(arr6)):
		for x in range(len(arr6[y])):
			arr6[y][x] = alternatif_karsilik_ucgensel_tup(arr6[y][x])
	
	all_tups.append(arr5)
	all_tups_copras.append(arr6)

	if i == 0:
		ws.title = "Katılımcı 1"

	else:
		wb.create_sheet("Katılımcı {}".format(i+1))
		
		wb.active = wb["Katılımcı {}".format(i+1)]

	ws = wb.active

	for k in range(3):
		for row in range(1, kriter_sayisi+1):
			cell = ws.cell(row=(k*(kriter_sayisi+3))+row+2, column=2)
			cell.value="K{}".format(row)

		for col in range(1, cozum_sayisi+1):
			cell = ws.cell(row=(k*(kriter_sayisi+3))+2, column=col+2)
			cell.value="Ç{}".format(col)

		for row in range(1, kriter_sayisi+1):
			for col in range(1, cozum_sayisi+1):
				cell = ws.cell(row=(k*(kriter_sayisi+3))+row+2, column=col+2)
				if k == 0:
					#print(arr2[row-1][col-1])
					cell.value = arr2[row-1][col-1]
				if k == 1:
					cell.value = arr3[row-1][col-1]
				if k == 2:
					cell.value = arr4[row-1][col-1]
				#cell.value=5

	# for row in range(1, kriter_sayisi+1):
	# 	cell = ws.cell(row=row+2, column=2)
	# 	cell.value="K{}".format(row)

	# for col in range(1, cozum_sayisi+1):
	# 	cell = ws.cell(row=2, column=col+2)
	# 	cell.value="Ç{}".format(col)


toplam_matris = []
averages = []
pda = []
nda = []
sp = []
np = []
nsp = []
nsn = []
asj = []
kas = []

copras = []
bnp = []
copras_normalized = []
copras_weighted = []
copras_si = []
copras_qi = []
copras_pi = []


agirliklar = [0.215
,0.207
,0.208
,0.183
,0.186
,0.338
,0.331
,0.331
,0.331
,0.327
,0.342
,0.207
,0.225
,0.182
,0.195
,0.192
,0.170
,0.176
,0.173
,0.180
,0.150
,0.151]

onem_derecesi = [0.045,
0.043,
0.043,
0.038,
0.039,
0.069,
0.068,
0.068,
0.066,
0.066,
0.068,
0.041,
0.045,
0.036,
0.039,
0.038,
0.031,
0.033,
0.032,
0.033,
0.028,
0.028
]


for row in range(cozum_sayisi):
	copras.append([])
	bnp.append([])
	copras_normalized.append([])
	copras_weighted.append([])
	for col in range(kriter_sayisi):
		copras[row].append((0,0,0))
		bnp[row].append(0)
		copras_normalized[row].append(0)
		copras_weighted[row].append(0)

for row in range(kriter_sayisi):
	toplam_matris.append([])
	pda.append([])
	nda.append([])
	for col in range(cozum_sayisi):
		toplam_matris[row].append((0,0,0))
		pda[row].append((0,0,0))
		nda[row].append((0,0,0))

for col in range(cozum_sayisi):
	sp.append((0,0,0))
	np.append((0,0,0))
	nsp.append((0,0,0))
	nsn.append((0,0,0))
	asj.append((0,0,0))
	kas.append(0)
	copras_pi.append(0)
	copras_qi.append(0)
	copras_si.append(0)

for i in range(len(all_tups)):
	for row in range(len(all_tups[i])):
		for col in range(len(all_tups[i][row])):
			toplam_matris[row][col] = ucgensel_toplam(toplam_matris[row][col],all_tups[i][row][col])

#print(toplam_matris)

for row in range(len(toplam_matris)):
	for col in range(len(toplam_matris[row])):
		toplam_matris[row][col] = ucgensel_bol_sabit(toplam_matris[row][col], katilimci_sayisi)

for row in range(len(toplam_matris)):
	toplam = (0, 0, 0)
	for col in range(len(toplam_matris[row])):
		toplam = ucgensel_toplam(toplam_matris[row][col], toplam)

	averages.append(ucgensel_bol_sabit(toplam, cozum_sayisi))

for row in range(len(toplam_matris)):
	for col in range(len(toplam_matris[row])):
		pda[row][col] = ucgensel_bol_sabit(psi(ucgensel_cik(toplam_matris[row][col], averages[row])), kav(averages[row]))
		nda[row][col] = ucgensel_bol_sabit(psi(ucgensel_cik(averages[row], toplam_matris[row][col])), kav(averages[row]))

#print(toplam_matris)
#print(averages)
#print(nda)
#print("********PDA***********")
#print(pda)

# print("*********NDA***********")
# print(nda)

for col in range(cozum_sayisi):
	for row in range(kriter_sayisi):
		sp[col] = ucgensel_toplam(sp[col], ucgensel_carp_sabit(pda[row][col], agirliklar[row]))
		np[col] = ucgensel_toplam(np[col], ucgensel_carp_sabit(nda[row][col], agirliklar[row]))

		# sp[col] = ucgensel_carp_sabit(ucgensel_toplam(sp[col], pda[row][col]), agirliklar[row])
		# np[col] = ucgensel_carp_sabit(ucgensel_toplam(np[col], nda[row][col]), agirliklar[row])


sp_maks = max(sp)
np_maks = max(np)

for col in range(cozum_sayisi):
	nsp[col] = ucgensel_bol_sabit(sp[col], sp_maks)
	nsn[col] = ucgensel_cik_sabit_rev(ucgensel_bol_sabit(np[col], np_maks), 1)

for col in range(cozum_sayisi):
	asj[col] = ucgensel_bol_sabit(ucgensel_toplam(nsp[col], nsn[col]), 2)

for col in range(cozum_sayisi):
	kas[col] = kav(asj[col])

wb.create_sheet("Bulanık EDAS")
wb.active = wb["Bulanık EDAS"]
ws = wb.active



for k in range(3):
	for row in range(1, kriter_sayisi+1):
		cell = ws.cell(row=(k*(kriter_sayisi+3))+row+2, column=2)
		cell.value="K{}".format(row)

	for col in range(1, cozum_sayisi+1):
		cell = ws.cell(row=(k*(kriter_sayisi+3))+2, column=col+2)
		cell.value="Ç{}".format(col)

	if k == 0:
		cell = ws.cell(row=2, column=cozum_sayisi+3)
		cell.value="AV"
		for row in range(1, kriter_sayisi+1):
			cell = ws.cell(row=row+2, column=cozum_sayisi+3)
			cell.value = tup_to_str(averages[row-1])

	for row in range(1, kriter_sayisi+1):
			for col in range(1, cozum_sayisi+1):
				cell = ws.cell(row=(k*(kriter_sayisi+3))+row+2, column=col+2)

				if k == 0:
					cell.value = tup_to_str(toplam_matris[row-1][col-1])

				if k == 1:
					cell.value = tup_to_str(pda[row-1][col-1])

				if k == 2:
					cell.value = tup_to_str(nda[row-1][col-1])


for row in range(1, cozum_sayisi+1):
	cell = ws.cell(row=(3*(kriter_sayisi+3))+row+2, column=2)
	cell.value="Ç{}".format(row)

for col in range(1, 6+1):
	cell = ws.cell(row=(3*(kriter_sayisi+3))+2, column=col+2)
	cell.value="s{}".format(col)

for col in range(1, cozum_sayisi+1):
	cell = ws.cell(row=(3*(kriter_sayisi+3))+row+2, column=col+2)

for row in range(1, cozum_sayisi+1):
	for col in range(1, 6+1):
		cell = ws.cell(row=(3*(kriter_sayisi+3))+row+2, column=col+2)
		if col == 1:
			cell.value = tup_to_str(sp[row-1])
		if col == 2:
			cell.value = tup_to_str(np[row-1])
		if col == 3:
			cell.value = tup_to_str(nsp[row-1])
		if col == 4:
			cell.value = tup_to_str(nsn[row-1])
		if col == 5:
			cell.value = tup_to_str(asj[row-1])
		if col == 6:
			cell.value = "{:.2f}".format(kas[row-1])

		


wb.create_sheet("Bulanık COPRAS")
wb.active = wb["Bulanık COPRAS"]
ws = wb.active

copras_i(all_tups_copras, kriter_sayisi, cozum_sayisi, katilimci_sayisi, copras)
copras_j(all_tups_copras, kriter_sayisi, cozum_sayisi, katilimci_sayisi, copras)
copras_k(all_tups_copras, kriter_sayisi, cozum_sayisi, katilimci_sayisi, copras)

for row in range(cozum_sayisi):
	for col in range(kriter_sayisi):
		cl = copras[row][col][0]
		cm = copras[row][col][1]
		cn = copras[row][col][2]
		bnp[row][col] = (((cn - cl) + (cm - cl)) / 3) + cl

for row in range(cozum_sayisi):
	for col in range(kriter_sayisi):
		summ = 0
		for row2 in range(cozum_sayisi):
			summ += bnp[row2][col]

		copras_normalized[row][col] = (bnp[row][col] / summ)
		copras_weighted[row][col] = copras_normalized[row][col] * onem_derecesi[col]

for row in range(cozum_sayisi):
	summ = 0
	for col in range(kriter_sayisi):
		summ += copras_weighted[row][col]

	copras_si[row] = summ
	copras_qi[row] = summ


q_max = max2(copras_qi)

for row in range(cozum_sayisi):
	copras_pi[row] = (copras_qi[row] / q_max) * 100


for k in range(3):
	for row in range(1, cozum_sayisi+1):
		cell = ws.cell(row=(k*(cozum_sayisi+3))+row+2, column=2)
		cell.value="Ç{}".format(row)

	for col in range(1, kriter_sayisi+1):
		cell = ws.cell(row=(k*(cozum_sayisi+3))+2, column=col+2)
		cell.value="K{}".format(col)


	for row in range(1, cozum_sayisi+1):
			for col in range(1, kriter_sayisi+1):
				cell = ws.cell(row=(k*(cozum_sayisi+3))+row+2, column=col+2)
				
				if k == 0:
					cell.value = tup_to_str(copras[row-1][col-1])

				if k == 1:
					cell.value = bnp[row-1][col-1]

				if k == 2:
					cell.value = copras_weighted[row-1][col-1]

for row in range(1, cozum_sayisi+1):
	cell = ws.cell(row=(3*(cozum_sayisi+3))+row+2, column=2)
	cell.value="Ç{}".format(row)

for col in range(1, 3+1):
	cell = ws.cell(row=(3*(cozum_sayisi+3))+2, column=col+2)
	cell.value="s{}".format(col)

for row in range(1, cozum_sayisi+1):
	for col in range(1, 3+1):
		cell = ws.cell(row=(3*(cozum_sayisi+3))+row+2, column=col+2)
		if col == 1:
			cell.value = copras_si[row-1]
		if col == 2:
			cell.value = copras_qi[row-1]
		if col == 3:
			cell.value = copras_pi[row-1]


wb.save("newsample.xlsx")

#s = pyperclip.paste()

# cozum_sayisi = int(input("Çözüm Sayısı :"))
# kriter_sayisi = int(input("Kriter Sayısı :"))
# katilimci_sayisi = int(input("Katılımcı Sayısı :"))

# print(cozum_sayisi, kriter_sayisi, katilimci_sayisi)
# bulanik_dizi = np.zeros((kriter_sayisi, cozum_sayisi))
# print(bulanik_dizi)


# lines = []
# while True:
#     line = input()
#     if line:
#         lines.append(line.replace('\t', ','))
#     else:
#         break

# text = '\n'.join(lines)
# test1 = text.split("\n")
# b = []
# for i in range(len(test1)):
# 	a = test1[i].split(',')
# 	lst_int = [int(x) for x in test1[i].split(',')]
# 	b.append(lst_int)

# #print(b)

# arr = np.array(b)

# arr2 = np.transpose(arr).tolist()

# print(arr2)

# for y in range(len(arr2)):
# 	for x in range(len(arr2[y])):
# 		arr2[y][x] = alternatif_karsilik_sozel(arr2[y][x])

# print(arr2)

	
